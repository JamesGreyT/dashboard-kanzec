"""Generic xlsx streaming export used by every ranked analytics endpoint.

Takes a list of column-specs + the rows (as dicts) and returns a
FastAPI `StreamingResponse`. Header row bold on a pale fill; numeric
columns right-aligned with Excel-native number formats; column widths
sized per column kind."""
from __future__ import annotations

import io
from dataclasses import dataclass
from datetime import date, datetime
from typing import Any, Iterable, Literal

from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ColKind = Literal["text", "int", "money", "pct", "date", "qty"]


@dataclass(frozen=True)
class ExportColumn:
    key: str
    label: str
    kind: ColKind = "text"
    width: int | None = None


def _width(c: ExportColumn) -> int:
    if c.width:
        return c.width
    return {"text": 22, "int": 12, "money": 16, "pct": 10, "date": 14, "qty": 12}[c.kind]


_NUMBER_FORMAT: dict[ColKind, str] = {
    "int":   "#,##0",
    "qty":   "#,##0.##",
    "money": '"$"#,##0.00',
    "pct":   "0.0%",
    "date":  "yyyy-mm-dd",
}


def _coerce(v: Any, kind: ColKind) -> Any:
    if v is None or v == "":
        return None
    if kind == "date":
        if isinstance(v, (date, datetime)):
            return v
        if isinstance(v, str):
            try:
                return date.fromisoformat(v[:10])
            except ValueError:
                return v
        return v
    if kind == "int":
        try:
            return int(v)
        except (TypeError, ValueError):
            return v
    if kind in ("money", "qty", "pct"):
        try:
            return float(v)
        except (TypeError, ValueError):
            return v
    return v


def stream_xlsx(
    *,
    filename: str,
    sheet_title: str,
    columns: list[ExportColumn],
    rows: Iterable[dict[str, Any]],
    totals: dict[str, Any] | None = None,
) -> StreamingResponse:
    wb = Workbook()
    ws = wb.active
    ws.title = (sheet_title[:31] or "Sheet1").replace("/", "-")

    header_fill = PatternFill("solid", fgColor="E7E3D5")
    header_font = Font(bold=True)
    totals_fill = PatternFill("solid", fgColor="F3F4F6")
    totals_font = Font(bold=True)
    right = Alignment(horizontal="right", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    # Column widths
    for i, c in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(i)].width = _width(c)

    # Header row
    for col, c in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col, value=c.label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = left

    # Data rows
    r = 2
    for row in rows:
        for col, c in enumerate(columns, 1):
            v = _coerce(row.get(c.key), c.kind)
            cell = ws.cell(row=r, column=col, value=v)
            fmt = _NUMBER_FORMAT.get(c.kind)
            if fmt:
                cell.number_format = fmt
            cell.alignment = right if c.kind in ("int", "money", "pct", "qty") else left
        r += 1

    # Totals row
    if totals:
        for col, c in enumerate(columns, 1):
            if col == 1:
                cell = ws.cell(row=r, column=col, value="TOTAL")
            else:
                v = _coerce(totals.get(c.key), c.kind)
                cell = ws.cell(row=r, column=col, value=v)
            cell.fill = totals_fill
            cell.font = totals_font
            cell.alignment = right if c.kind in ("int", "money", "pct", "qty") else left
            fmt = _NUMBER_FORMAT.get(c.kind)
            if fmt and col > 1:
                cell.number_format = fmt

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    ts = datetime.now().strftime("%Y%m%d-%H%M")
    download_name = f"{filename}-{ts}.xlsx"
    return StreamingResponse(
        buf,
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={"Content-Disposition": f'attachment; filename="{download_name}"'},
    )
