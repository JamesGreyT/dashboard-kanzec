"""SQL builder for the DataViewer — whitelist-driven.

A filter coming in as ?filter[delivery_date]=>=:2026-04-01 parses to a
(column, op, value) triple. Both column name AND operator are checked
against the TableDef in catalog.py; value is always passed as a bound
parameter. If anything falls outside the catalog, we raise 400 up the stack.
"""
from __future__ import annotations

import csv
import io
from datetime import date, datetime
from decimal import Decimal
from typing import Any, AsyncIterator

from fastapi import HTTPException, status
from sqlalchemy import bindparam, text
from sqlalchemy.ext.asyncio import AsyncSession

from ..scope import UserScope, clause_for_table
from .catalog import CATALOG, ColumnDef, TableDef


def _table(key: str) -> TableDef:
    t = CATALOG.get(key)
    if t is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, f"unknown table {key!r}")
    return t


def _parse_filter_value(col: ColumnDef, raw: str) -> Any:
    """Coerce a string from the URL into the column's type. 400 on bad input."""
    try:
        if col.type == "date":
            return date.fromisoformat(raw)
        if col.type == "timestamp":
            # Accept either plain ISO date or full ISO datetime.
            if "T" in raw or " " in raw:
                return datetime.fromisoformat(raw.replace(" ", "T"))
            return date.fromisoformat(raw)
        if col.type == "int":
            return int(raw)
        if col.type == "numeric":
            return Decimal(raw)
        return raw  # text
    except (ValueError, TypeError) as e:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, f"bad value for {col.label}: {e}") from None


def _build_where(
    table: TableDef,
    filters: list[tuple[str, str, str]],
    search: str | None,
    scope: UserScope | None = None,
) -> tuple[str, dict[str, Any]]:
    """Returns (where_clause_sql, params_dict). Always safe — no user strings
    land in the SQL itself, only catalog-whitelisted column names and ops.

    `filters` is a list of (column, op, value_str) triples. Multiple filters
    on the same column are ANDed (e.g. `>= 2026-04-01` + `<= 2026-04-30`).

    If `scope` is set and the user is scoped, a per-table scope fragment is
    AND'd in after the caller's filters.
    """
    clauses: list[str] = []
    params: dict[str, Any] = {}

    for i, (col_name, op, value_str) in enumerate(filters):
        col = table.columns.get(col_name)
        if col is None:
            raise HTTPException(status.HTTP_400_BAD_REQUEST, f"unknown column {col_name!r}")
        if op not in col.ops:
            raise HTTPException(
                status.HTTP_400_BAD_REQUEST,
                f"operator {op!r} not allowed for column {col_name}",
            )
        pname = f"f{i}"
        if op == "in":
            # value is a pipe-separated list: "USD|UZS|EUR"
            raw_values = [v for v in value_str.split("|") if v != ""]
            if not raw_values:
                # empty IN → no match; keep the SQL valid by matching nothing.
                clauses.append("false")
                continue
            parsed = [_parse_filter_value(col, v) for v in raw_values]
            placeholders = ", ".join(f":{pname}_{j}" for j in range(len(parsed)))
            clauses.append(f'"{col_name}" IN ({placeholders})')
            for j, v in enumerate(parsed):
                params[f"{pname}_{j}"] = v
        elif op == "ilike":
            value = _parse_filter_value(col, value_str)
            clauses.append(f'"{col_name}"::text ILIKE :{pname}')
            params[pname] = f"%{value}%"
        else:
            value = _parse_filter_value(col, value_str)
            clauses.append(f'"{col_name}" {op} :{pname}')
            params[pname] = value

    if search:
        # Search across text-ish columns only.
        textish = [
            name for name, c in table.columns.items()
            if c.type == "text" or c.id_column
        ]
        if textish:
            or_parts = []
            for i, name in enumerate(textish):
                pname = f"s{i}"
                or_parts.append(f'"{name}"::text ILIKE :{pname}')
                params[pname] = f"%{search}%"
            clauses.append("(" + " OR ".join(or_parts) + ")")

    if scope is not None:
        frag, scope_params = clause_for_table(
            scope, f"{table.schema}.{table.table}"
        )
        if frag:
            clauses.append(frag)
            params.update(scope_params)

    where = ("WHERE " + " AND ".join(clauses)) if clauses else ""
    return where, params


def _build_order_by(table: TableDef, sort_raw: str | None) -> str:
    """sort_raw shaped like 'delivery_date:desc,product_amount:asc'."""
    if not sort_raw:
        parts = [(f, d) for f, d in table.default_sort]
    else:
        parts = []
        for chunk in sort_raw.split(","):
            if not chunk.strip():
                continue
            f, _, d = chunk.partition(":")
            f = f.strip()
            d = d.strip().lower() or "asc"
            if f not in table.columns:
                raise HTTPException(status.HTTP_400_BAD_REQUEST, f"unknown sort column {f!r}")
            if d not in ("asc", "desc"):
                raise HTTPException(status.HTTP_400_BAD_REQUEST, f"bad sort direction {d!r}")
            parts.append((f, d))
    if not parts:
        return ""
    return "ORDER BY " + ", ".join(f'"{f}" {d} NULLS LAST' for f, d in parts)


# ---- Public ----------------------------------------------------------------

async def list_rows(
    session: AsyncSession,
    key: str,
    *,
    filters: list[tuple[str, str, str]],
    search: str | None,
    sort: str | None,
    limit: int,
    offset: int,
    scope: UserScope | None = None,
) -> dict[str, Any]:
    table = _table(key)
    where, params = _build_where(table, filters, search, scope=scope)
    order_by = _build_order_by(table, sort)

    col_list = ", ".join(f'"{c}"' for c in table.columns)
    base = f'FROM "{table.schema}"."{table.table}" {where}'

    count_row = (
        await session.execute(text(f"SELECT COUNT(*) AS c {base}"), params)
    ).one()
    total = int(count_row.c)

    params_paged = {**params, "lim": limit, "off": offset}
    rows = (
        await session.execute(
            text(f"SELECT {col_list} {base} {order_by} LIMIT :lim OFFSET :off"),
            params_paged,
        )
    ).mappings().all()

    return {
        "rows": [_jsonify(r) for r in rows],
        "total": total,
        "limit": limit,
        "offset": offset,
    }


async def distinct_values(
    session: AsyncSession,
    key: str,
    column: str,
    *,
    search: str | None,
    limit: int,
    scope: UserScope | None = None,
) -> dict[str, Any]:
    """Return up to `limit` distinct values of `column`, with counts, ordered
    by frequency (desc) then value (asc). Powers the Excel-style checkbox
    filter."""
    table = _table(key)
    col = table.columns.get(column)
    if col is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, f"unknown column {column!r}")
    # Cap hard so a rogue client can't pull 1M distincts.
    limit = max(1, min(limit, 500))

    params: dict[str, Any] = {"lim": limit + 1}
    where_parts: list[str] = []
    if search:
        where_parts.append(f'"{column}"::text ILIKE :q')
        params["q"] = f"%{search}%"
    if scope is not None:
        frag, scope_params = clause_for_table(scope, f"{table.schema}.{table.table}")
        if frag:
            where_parts.append(frag)
            params.update(scope_params)
    where = ("WHERE " + " AND ".join(where_parts)) if where_parts else ""

    rows = (
        await session.execute(
            text(f"""
                SELECT "{column}"::text AS v, COUNT(*) AS c
                  FROM "{table.schema}"."{table.table}"
                  {where}
                 GROUP BY 1
                 ORDER BY c DESC NULLS LAST, v ASC NULLS LAST
                 LIMIT :lim
            """),
            params,
        )
    ).all()
    limited = len(rows) > limit
    rows = rows[:limit]
    return {
        "values": [{"value": r.v, "count": int(r.c)} for r in rows],
        "limited": limited,
    }


async def get_row(
    session: AsyncSession,
    key: str,
    pk_values: list[str],
    *,
    scope: UserScope | None = None,
) -> dict:
    table = _table(key)
    if len(pk_values) != len(table.pk):
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            f"expected {len(table.pk)} pk parts, got {len(pk_values)}",
        )
    where_parts = []
    params: dict[str, Any] = {}
    for i, (name, raw) in enumerate(zip(table.pk, pk_values)):
        col = table.columns[name]
        pname = f"p{i}"
        where_parts.append(f'"{name}" = :{pname}')
        params[pname] = _parse_filter_value(col, raw)
    if scope is not None:
        frag, scope_params = clause_for_table(scope, f"{table.schema}.{table.table}")
        if frag:
            where_parts.append(frag)
            params.update(scope_params)
    col_list = ", ".join(f'"{c}"' for c in table.columns)
    row = (
        await session.execute(
            text(f'SELECT {col_list} FROM "{table.schema}"."{table.table}" WHERE ' + " AND ".join(where_parts)),
            params,
        )
    ).mappings().first()
    if row is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "row not found")
    return _jsonify(row)


async def stream_csv(
    session: AsyncSession,
    key: str,
    *,
    filters: list[tuple[str, str, str]],
    search: str | None,
    sort: str | None,
    max_rows: int = 100_000,
    scope: UserScope | None = None,
) -> AsyncIterator[bytes]:
    """Yield CSV bytes in batches. Bounded at max_rows to avoid runaway exports."""
    table = _table(key)
    where, params = _build_where(table, filters, search, scope=scope)
    order_by = _build_order_by(table, sort)
    col_list = ", ".join(f'"{c}"' for c in table.columns)

    # Header
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(list(table.columns.keys()))
    yield buf.getvalue().encode("utf-8")

    params_paged = {**params, "lim": max_rows, "off": 0}
    stream = await session.stream(
        text(f'SELECT {col_list} FROM "{table.schema}"."{table.table}" {where} {order_by} LIMIT :lim OFFSET :off'),
        params_paged,
    )
    async for chunk in stream.mappings().partitions(1000):
        buf = io.StringIO()
        writer = csv.writer(buf)
        for r in chunk:
            writer.writerow([_csv_cell(r[c]) for c in table.columns])
        yield buf.getvalue().encode("utf-8")


async def build_xlsx(
    session: AsyncSession,
    key: str,
    *,
    filters: list[tuple[str, str, str]],
    search: str | None,
    sort: str | None,
    max_rows: int = 100_000,
    scope: UserScope | None = None,
) -> bytes:
    """Build an xlsx workbook honoring the current filter/sort/search state.

    openpyxl's write-only mode keeps per-row memory low; the final file is
    packaged into ZIP bytes at save time (xlsx can't be truly streamed — the
    central directory lands last). Header row uses column labels from the
    catalog so the export reads like the dashboard, not raw SQL identifiers.
    """
    from openpyxl import Workbook
    from openpyxl.cell import WriteOnlyCell
    from openpyxl.styles import Font, PatternFill, Alignment

    table = _table(key)
    where, params = _build_where(table, filters, search, scope=scope)
    order_by = _build_order_by(table, sort)
    col_list = ", ".join(f'"{c}"' for c in table.columns)

    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title=table.label[:31] or key[:31])

    # Header row — bold on a muted fill so it reads as a header in Excel.
    header_font = Font(bold=True, color="1F2937")
    header_fill = PatternFill("solid", fgColor="F3F4F6")
    center = Alignment(horizontal="left", vertical="center")
    header_cells = []
    for col_name, col_def in table.columns.items():
        c = WriteOnlyCell(ws, value=col_def.label or col_name)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        header_cells.append(c)
    ws.append(header_cells)

    # Column widths — a rough heuristic so dates don't hug and text isn't huge.
    widths = []
    for col_name, col_def in table.columns.items():
        if col_def.type in ("date",):
            widths.append(12)
        elif col_def.type in ("timestamp",):
            widths.append(18)
        elif col_def.type in ("int", "numeric"):
            widths.append(14)
        elif col_def.id_column:
            widths.append(16)
        else:
            widths.append(24)
    ws.column_dimensions  # touch to ensure dict exists
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[_col_letter(i)].width = w

    params_paged = {**params, "lim": max_rows, "off": 0}
    stream = await session.stream(
        text(
            f'SELECT {col_list} FROM "{table.schema}"."{table.table}" '
            f'{where} {order_by} LIMIT :lim OFFSET :off'
        ),
        params_paged,
    )
    async for chunk in stream.mappings().partitions(1000):
        for r in chunk:
            ws.append([_xlsx_cell(r[c]) for c in table.columns])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _col_letter(n: int) -> str:
    # 1 -> "A", 27 -> "AA" — enough for any sane column count.
    s = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def _xlsx_cell(v):
    """Coerce a DB value to something openpyxl writes sensibly.
    Decimal → float (Excel has no Decimal). tz-aware datetimes are
    converted to naive (Excel has no timezones — openpyxl raises
    TypeError on tzinfo). date/datetime pass through so Excel
    recognises them as dates. Strings/ints/None pass through."""
    if v is None:
        return None
    if isinstance(v, Decimal):
        return float(v)
    if isinstance(v, datetime) and v.tzinfo is not None:
        return v.replace(tzinfo=None)
    if isinstance(v, (date, datetime)):
        return v
    return v


def _jsonify(row) -> dict:
    out = {}
    for k, v in row.items():
        if isinstance(v, Decimal):
            out[k] = float(v)
        elif isinstance(v, datetime):
            out[k] = v.isoformat()
        elif isinstance(v, date):
            out[k] = v.isoformat()
        else:
            out[k] = v
    return out


def _csv_cell(v) -> str:
    if v is None:
        return ""
    if isinstance(v, Decimal):
        return str(float(v))
    if isinstance(v, (date, datetime)):
        return v.isoformat()
    return str(v)
